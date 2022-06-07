import datetime
import glob
import itertools
import math as math
import os
import pickle
from natsort import natsorted

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import torch
from PIL import Image
from matplotlib import cm
from sklearn.metrics import auc
from sklearn.metrics import confusion_matrix
from sklearn.metrics import roc_curve
from matplotlib.ticker import LinearLocator, FormatStrFormatter


### PSO discovery ###
def plot_pso_convergence(plot_training_label_dir, global_best_val):
    fig = plt.figure()
    plt.plot(global_best_val)
    plt.xlabel('Number of Iterations')
    plt.ylabel('Global Best Value')
    fig.savefig(os.path.join(plot_training_label_dir, f"pso_iter.png"), dpi=400, format='png')
    # plt.show()

def get_cmap(n, name='hsv'):
    """Returns a function that maps each index in 0, 1, ..., n-1 to a distinct
    RGB color; the keyword argument name must be a standard mpl colormap name."""
    return plt.cm.get_cmap(name, n)

def plot_features_last_iteration(plot_training_dir, history_particles, dim_space):
    plt.rcParams['figure.figsize'] = [8, 6]
    plt.rcParams['figure.dpi'] = 400
    plt.rcParams['font.size'] = 12

    fig = plt.figure()
    cmap = get_cmap(dim_space)
    for dim in np.arange(dim_space):
        for p_key in history_particles.keys():
            final_p_pos_particle_dim = history_particles[p_key].iloc[-1, dim]
            plt.scatter(final_p_pos_particle_dim, dim, s=10.0, marker="o", edgecolors="None", alpha=1,  c=[cmap(dim)])
    plt.gca().xaxis.grid(True)
    plt.xlabel("Particles Position")
    plt.ylabel("Dimensions")
    plt.title("Particle Position fo each dimension at last PSO iteration")
    fig.savefig(os.path.join(plot_training_dir, f"pso_dim_last_iteration.png"), dpi=400, format='png')
    #plt.show()

def plot_feature(plot_training_dir, history_particles, dim_space, iteration):
    plt.rcParams['figure.figsize'] = [8, 6]  # [12, 6] # default = [6.0, 4.0]
    plt.rcParams['figure.dpi'] = 400  # default = 72.0
    plt.rcParams['font.size'] = 12  # default = 10.0

    iterations = np.arange(iteration)
    alphas = np.linspace(0.1, 0.5, num=iteration)
    cmap = get_cmap(dim_space)
    for idx, dimension_to_plot in enumerate(range(dim_space)): # cycle across dimension
        print(f"dim:{dimension_to_plot}")
        fig = plt.figure()
        for p_key in history_particles.keys():  # cycle across particle
            history_particle = history_particles[p_key]
            history_particle_dim = history_particle.iloc[:, dimension_to_plot]
            plt.scatter(history_particle_dim, iterations, marker="o", edgecolors="None", alpha=alphas, c=[cmap(idx)])
            plt.xlabel(f'Dimension {dimension_to_plot}')
            plt.ylabel(f'Iteration')
            plt.title(f'Particles position Dimension {dimension_to_plot} across Iterations')
        fig.savefig(os.path.join(plot_training_dir, f"pso_dim_{dimension_to_plot}.png"), dpi=400, format='png')
        #plt.show()

def make_gif_from_folder(plot_training_dir, img_to_upload_dir, filename_result='iid_img.gif', filename_source='pso_images'):
    files = os.path.join(img_to_upload_dir, f'{filename_source}_*.png')
    sorted_files = natsorted(glob.glob(files))
    img, *imgs = [Image.open(f) for f in sorted_files]
    img.save(fp=os.path.join(plot_training_dir, filename_result), format='GIF', append_images=imgs, save_all=True,  duration=200, loop=0)


def plot2d(plot_training_dir, history_particles, fitness, g_best_pos, grid_range=5, color = '#000'):

    assert len(history_particles['particle_0'].iloc[0, :]) == 2
    assert isinstance(history_particles, dict)

    n_particles = len(history_particles)
    n_iterations = len(history_particles['particle_0'])

    plt.rcParams['figure.figsize'] = [8, 6]
    plt.rcParams['figure.dpi'] = 400
    plt.rcParams['font.size'] = 12
    cmap = cm.colors.LinearSegmentedColormap.from_list('Custom', [(0, '#2f9599'), (0.45, '#eee'), (1, '#8800ff')],  N=256)

    # Define a coordinate grid using the best position finded by the Swarm
    X = np.arange(g_best_pos[0] - grid_range, g_best_pos[0] + grid_range, 0.1)
    Y = np.arange(g_best_pos[1] - grid_range, g_best_pos[1] + grid_range, 0.1)
    meshgrid = np.meshgrid(X, Y)

    # Compute the fitness for each point in the grid
    print('Compute the meshgrid')
    X_grid, Y_grid = meshgrid
    meshdim = len(meshgrid[0])
    Z_grid = np.zeros((meshdim, meshdim), dtype=np.float64)
    img_grid = []
    for i, (x, y) in enumerate(zip(X_grid, Y_grid)):
        for j, (xx, yy) in enumerate(zip(x, y)):
            f, im = fitness(dim_space=2, pos=np.array([xx, yy], dtype=np.float64))
            Z_grid[i, j] = f.item()
            img_grid.append(im)

    # One plot for iteration
    for iteration in range(0, n_iterations):
        # Get coordinates for all particles at the current iteration
        data_iid = np.ones((n_particles, 2), dtype='float32')
        for particle_idx, p_key in enumerate(history_particles.keys()):
            data_iid[particle_idx, :] = history_particles[p_key].iloc[iteration, :]

        X = data_iid[:, 0]
        Y = data_iid[:, 1]
        # Z = [fitness(dim_space=2, pos=np.array([x, y], dtype=np.float64))[0].item() for x, y in zip(X, Y)]

        # Plot
        fig, ax = plt.subplots()
        ax.contour(X_grid, Y_grid, Z_grid, levels=30, linewidths=0.5, colors='#999') # Add contours and contours lines
        cntr = ax.contourf(X_grid, Y_grid, Z_grid, levels=30, cmap=cmap, alpha=0.5)
        ax.scatter(X, Y, color=color)
        ax.scatter(g_best_pos[0], g_best_pos[1], s=50, marker='x', color='red', alpha=1)

        # Add labels and set equal aspect ratio
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_xlim(np.min(X_grid), np.max(X_grid)) # todo compute the min max values considering the whole range during iterations
        ax.set_ylim(np.min(Y_grid), np.max(Y_grid))
        ax.set_aspect(aspect='equal')
        plt.colorbar(cntr, ax=ax)
        plt.title(f'Iteration: {iteration}')
        plt.savefig(os.path.join(plot_training_dir, f"2d_plot_{iteration}.png"), dpi=400, format='png')
        #plt.show()

    return Z_grid, img_grid

def plot_training(history, plot_training_dir, label=None):
    if not isinstance(history, pd.DataFrame):
        history = pd.DataFrame.from_dict(history, orient='index').transpose()
    if 'train_loss' in history.columns and 'val_loss' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_loss', 'val_loss'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title("Training and validation loss")
        plt.xlabel('Epochs')
        plt.ylabel('Average Negative Log Likelihood')
        plt.legend()
        if label is not None:
            plt.savefig(os.path.join(plot_training_dir, f"train_val_loss_{label}.png"), dpi=400, format='png')
        else:
            plt.savefig(os.path.join(plot_training_dir, f"train_val_loss.png"), dpi=400, format='png')
        plt.show()
    if 'train_acc' in history.columns and 'val_acc' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_acc', 'val_acc'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title('Training and Validation Accuracy')
        plt.xlabel('Epochs')
        plt.ylabel('Average Accuracy')
        plt.ylim([-0.1, 1.1])
        plt.legend()
        if label is not None:
            plt.savefig(os.path.join(plot_training_dir, f"train_val_acc_{label}.png"), dpi=400, format='png')
        else:
            plt.savefig(os.path.join(plot_training_dir, f"train_val_acc.png"), dpi=400, format='png')
        plt.show()
    if 'train_auc' in history.columns and 'val_auc' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_auc', 'val_auc'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title('Training and Validation AUC')
        plt.xlabel('Epochs')
        plt.ylabel('AUC')
        plt.ylim([-0.1, 1.1])
        plt.legend()
        plt.savefig(os.path.join(plot_training_dir, "train_val_auc.png"), dpi=400, format='png')
        plt.show()
    if 'train_f1' in history.columns and 'val_f1' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_f1', 'val_f1'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title('Training and Validation f1-score')
        plt.xlabel('Epochs')
        plt.ylabel('f1-score')
        plt.ylim([-0.1, 1.1])
        plt.legend()
        plt.savefig(os.path.join(plot_training_dir, "train_val_f1-score.png"), dpi=400, format='png')
        plt.show()
    if 'train_prec' in history.columns and 'val_prec' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_prec', 'val_prec'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title('Training and Validation Precision')
        plt.xlabel('Epochs')
        plt.ylabel('Precision')
        plt.ylim([-0.1, 1.1])
        plt.legend()
        plt.savefig(os.path.join(plot_training_dir, "train_val_precision.png"), dpi=400, format='png')
        plt.show()
    if 'train_rec' in history.columns and 'val_rec' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_rec', 'val_rec'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title('Training and Validation Recall')
        plt.xlabel('Epochs')
        plt.ylabel('Recall')
        plt.ylim([-0.1, 1.1])
        plt.legend()
        plt.savefig(os.path.join(plot_training_dir, "train_val_recall.png"), dpi=400, format='png')
        plt.show()
    if 'mean_mse' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['mean_mse'],['r']):
            plt.plot(history[c], label=c, color=color)
        plt.title('mse between particles position')
        plt.xlabel('Iterations')
        plt.ylabel('mean_mse')
        plt.savefig(os.path.join(plot_training_dir, "mean_mse.png"), dpi=400, format='png')
        # plt.show()

def write_readme(reports_dir, tag_name, filename, **attributes):
    """
    Function to write on .txt file a dictionary
    Args:
        reports_dir (String): folder to save data
        tag_name (String): header to include in .txt file
        filename
        attributes: attributes of dictionary
    """
    print('\n[INFO]', tag_name, sep=' ')
    [print(name, value, sep=': ') for name, value in attributes.items()]
    with open(os.path.join(reports_dir, filename), "a") as file:
        if tag_name[:] == '/nStart':
            file.write("START: " + datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
        file.write(tag_name)
        [file.write("\n{}: {}".format(name, value)) for name, value in attributes.items()]
        file.write("\n\n")

def append_df_to_excel(filename, dataframe, sheet_name='Sheet1', startrow=None,  **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename] into [sheet_name] Sheet. If [filename] doesn't exist, then this function will create it.
        Parameters:
            filename: File path or existing ExcelWriter (Example: '/path/to/file.xlsx')
            dataframe: dataframe to save to workbook
            sheet_name: Name of sheet which will contain DataFrame (default: 'Sheet1')
            startrow: upper left cell row to dump data frame. Per default (startrow=None) calculate the last row in the existing DF and write to the next row
            to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`  [can be dictionary]
        Returns:
             None
    """
    header = True
    if 'engine' in to_excel_kwargs: # ignore [engine] parameter if it was passed
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = openpyxl.load_workbook(filename)  # try to open an existing workbook
        if startrow is None and sheet_name in writer.book.sheetnames:  # get the last row in the existing Excel sheet if it was not specified explicitly
            startrow = writer.book[sheet_name].max_row
            header = False
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}  # copy existing sheets
    except FileNotFoundError:  # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    dataframe.to_excel(writer, sheet_name, startrow=startrow, header=header, **to_excel_kwargs)  # write out the new sheet
    writer.save()  # save the workbook


def write_to_excel(reports_dir, filename,  dictionary, selected_keys=None, fold=None, group="slices"):
    df = pd.DataFrame()
    if fold is None:
        fold = 999
    if selected_keys is None:
        selected_keys = list(dictionary.keys())
    if group[:] == 'patients_predictions':
        for key in selected_keys:
            df.loc[:, key] = np.asarray(dictionary[key])
    else:
        df.loc[fold, 'fold'] = fold
        for key in selected_keys:
            df.loc[fold, key] = dictionary[key]

    append_df_to_excel(filename=os.path.join(reports_dir, filename), dataframe=df, sheet_name=group, startrow=None, index=False)


def on_patients_prediction(patients_id_test, pred, id_test):
    pred_on_patients = [pred[id_test == idd] for idd in patients_id_test]
    return np.asarray([np.mean(x) for x in pred_on_patients])


def on_patients_majority_votes(patients_id_test, pred, id_test):
    pred_on_patients = [pred[id_test == idd] for idd in patients_id_test]
    n_pred_on_patients = np.asarray([(np.sum(x == 0), np.sum(x == 1)) for x in pred_on_patients])
    return np.argmax(n_pred_on_patients, axis=1)


def compute_metrics(dictionary, label, pred):
    # Conf matrix
    #    predicted
    #    TN     FP
    #
    #    FN     TP
    conf_mat = confusion_matrix(label, pred)
    tn, fp, fn, tp = conf_mat.ravel()
    dictionary['tn'], dictionary['fp'], dictionary['fn'], dictionary['tp'] = tn, fp, fn, tp
    dictionary['total_neg'] = tn + fp
    dictionary['total_pos'] = tp + fn

    dictionary['accuracy'] = (tp + tn) / (tp + tn + fp + fn)
    dictionary['recall'] = tp / (tp + fn)
    dictionary['precision'] = tp / (tp + fp)
    dictionary['f1'] = (2 * dictionary['recall'] * dictionary['precision']) /  (dictionary['recall'] + dictionary['precision'])
    dictionary['specificity'] = tn / (tn + fp)
    dictionary['g'] = math.sqrt(dictionary['recall'] * dictionary['specificity'])

    return dictionary


class Eval:
    def __init__(self):

        self.x = np.linspace(0, 1, 30)  # _points to consider for estimate the mean roc
        self.selected_keys = ['accuracy', 'precision', 'recall', 'f1', 'auc', 'specificity', 'g', 'tn', 'tp', 'fp', 'fn', 'total_neg', 'total_pos']
        self.overall_slices_score = []
        self.overall_patients_score = []
        self.overall_patients_predictions = {"ID_patients": [], "label": [], "majority_criteria": [], "mean_score": []}

    def evaluate_model(self, reports_dir, report_fold_dir, fold, model, data_loader, data, device):
        # Global and Class Accuracy
        outputs = []
        predictions = []
        y_test = []
        id_test = []

        # Test loop
        model.eval()
        with torch.no_grad():
            for x_batch, y_batch, id_batch, id_slices_batch in data_loader:
                x_batch = x_batch.to(device)
                y_batch = y_batch.to(device)

                output = model(x_batch.float()) # Prediction
                _, preds = torch.max(output, 1)
                outputs.append(output.cpu().detach().numpy())
                predictions.append(preds.cpu().detach().numpy())
                y_test.append(y_batch.cpu().detach().numpy())
                id_test.append(id_batch)

        outputs = np.asarray(list(itertools.chain(*outputs)))
        predictions = np.asarray(list(itertools.chain(*predictions)))
        y_test = np.asarray(list(itertools.chain(*y_test)))
        id_test = np.asarray(list(itertools.chain(*id_test)))

        patients_id_test = np.unique(data.index.values)
        patients_y_test = np.asarray([data.loc[idp, "label"][0] for idp in patients_id_test])


        for group in ["slices", "patients"]:
            score_dict = {}


            if group[:] == "slices":
                score_dict = compute_metrics(dictionary=score_dict, label=y_test, pred=predictions)
                score_dict = self.plot_roc(reports_dir=report_fold_dir, dictionary=score_dict, label=y_test, pred=outputs[:, 1], group=group)
                self.overall_slices_score.append(score_dict)
            else:
                binary_patients_predictions = on_patients_majority_votes(patients_id_test=patients_id_test, pred=predictions, id_test=id_test)
                patients_predictions = on_patients_prediction(patients_id_test=patients_id_test, pred=outputs[:, 1], id_test=id_test)
                score_dict = compute_metrics(dictionary=score_dict, label=patients_y_test, pred=binary_patients_predictions)
                score_dict = self.plot_roc(reports_dir=report_fold_dir, dictionary=score_dict, label=patients_y_test, pred=patients_predictions, group=group)
                self.overall_patients_score.append(score_dict)

                self.overall_patients_predictions["ID_patients"].append(patients_id_test)
                self.overall_patients_predictions["label"].append(patients_y_test)
                self.overall_patients_predictions["majority_criteria"].append(binary_patients_predictions)
                self.overall_patients_predictions["mean_score"].append(patients_predictions)

            write_readme(reports_dir=reports_dir, tag_name="FOLD" + str(fold) + group, filename='scores.txt',  **{key: score_dict[key] for key in self.selected_keys})
            write_to_excel(reports_dir=reports_dir, filename="results.xlsx", dictionary=score_dict, selected_keys=self.selected_keys, fold=fold, group=group)

    def plot_roc(self, reports_dir, dictionary, label, pred, group="slices"):
        fpr, tpr, thresholds = roc_curve(label, pred, pos_label=1)
        roc_auc = auc(fpr, tpr)

        plt.figure()
        plt.plot(fpr, tpr, color='black', label='AUC=' + str(roc_auc))
        plt.plot(self.x, self.x, 'k-')
        plt.legend()
        plt.xlabel('FP rate')
        plt.ylabel('TP rate')
        plt.savefig(os.path.join(reports_dir, "roc_curve_{}.png").format(group), dpi=400, format='png')
        plt.close("all")
        plt.clf()

        interp_tpr = np.interp(self.x, fpr, tpr)  # save data for plot_roc_mean
        interp_fpr = np.interp(self.x, tpr, fpr)
        interp_tpr[0] = 0.0
        interp_fpr[0] = 0.0

        dictionary['fprs'], dictionary['tprs'], dictionary['auc'] = interp_fpr, interp_tpr, roc_auc
        return dictionary

    def on_experiments_end(self, reports_dir, report_test_dir):
        with open(os.path.join(reports_dir, 'overall_scores.pkl'), 'wb') as f:
            pickle.dump([self.overall_slices_score, self.overall_patients_score], f)

        selected_keys = ['accuracy', 'precision', 'recall', 'f1', 'auc', 'specificity', 'g']
        std_keys = ['std_acc', 'std_prec', 'std_rec', 'std_f1', 'std_auc', 'std_spec', 'std_g']
        mean_dict_group = {"slices": [], "patients": []}
        std_dict_group = {"slices": [], "patients": []}
        for score_dict, group in zip([self.overall_slices_score, self.overall_patients_score], ["slices", "patients"]):
            mean_dict = {key: np.mean([x.get(key) for x in score_dict]) for key in selected_keys}
            std_dict = {key: np.std([x.get(key) for x in score_dict]) for key in selected_keys}
            write_readme(reports_dir=reports_dir, tag_name="MEAN_VALUE_"+group, filename='scores.txt', **mean_dict)
            write_to_excel(reports_dir=reports_dir, filename="results.xlsx", dictionary=mean_dict, selected_keys=selected_keys, fold='MEAN', group=group)
            write_to_excel(reports_dir=reports_dir, filename="results.xlsx", dictionary=std_dict,  selected_keys=selected_keys, fold='STD', group=group)
            self.mean_plot_roc(reports_dir = report_test_dir, dictionary_list= score_dict, group=group)

            for old_key, new_key in zip(selected_keys, std_keys):
                std_dict[new_key] = std_dict.pop(old_key)

            mean_dict_group[group] = mean_dict
            std_dict_group[group] = std_dict

        dictionary = \
            {'ID_patients': np.asarray(list(itertools.chain(*self.overall_patients_predictions['ID_patients']))),
             'label': np.asarray(list(itertools.chain(*self.overall_patients_predictions['label']))),
             'majority_criteria': np.asarray(list(itertools.chain(*self.overall_patients_predictions['majority_criteria']))),
             'mean_score': np.asarray(list(itertools.chain(*self.overall_patients_predictions['mean_score'])))}
        write_to_excel(reports_dir=reports_dir ,filename="results.xlsx", dictionary= dictionary, selected_keys= list(dictionary.keys()), fold='overall', group="patients_predictions")

        return mean_dict_group, std_dict_group

    def mean_plot_roc(self, reports_dir, dictionary_list, group):
        plt.figure()

        mean_tpr = np.mean([x.get('tprs') for x in dictionary_list], axis=0)
        mean_tpr[-1] = 1.0
        std_tpr = np.std([x.get('tprs') for x in dictionary_list], axis=0)

        std_fpr = np.std([x.get('fprs') for x in dictionary_list], axis=0)

        mean_auc = auc(self.x, mean_tpr)
        std_auc = np.std([x.get('auc') for x in dictionary_list])

        plt.plot([0, 1], [0, 1], linestyle='--', lw=2, color='gray', alpha=.8)

        plt.errorbar(self.x, mean_tpr, yerr=std_tpr, marker='s', capsize=5, capthick=2, elinewidth=2, ecolor='gray', fmt='-o', color='b', label=r'ROC media (AUC = %0.2f $\pm$ %0.2f)' % (mean_auc, std_auc), lw=2, alpha=.8)
        plt.errorbar(self.x, mean_tpr, xerr=std_fpr, marker='s', elinewidth=0.8, ecolor='gray', fmt='-o', color='b', lw=2, alpha=.8)

        plt.xlim([-0.05, 1.05])
        plt.ylim([-0.05, 1.05])
        plt.title(group + " mean roc curve", fontsize=14)
        plt.xlabel('FP Rate', fontsize=14)
        plt.ylabel('TP Rate', fontsize=14)
        plt.legend(loc="lower right", fontsize=12)

        plt.savefig(os.path.join(reports_dir, "mean_roc_" + group + ".png"), dpi=400, format='png')
        plt.close("all")
        plt.clf()


